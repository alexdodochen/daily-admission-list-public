"""
Migrate 每日入院名單.xlsx to Google Sheet
One-time migration script
"""
import sys, io, time, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
import gspread
from google.oauth2.service_account import Credentials
from gspread_formatting import *

# --- Auth ---
scopes = [
    'https://www.googleapis.com/auth/spreadsheets',
    'https://www.googleapis.com/auth/drive'
]
creds = Credentials.from_service_account_file(
    'sigma-sector-492215-d2-0612bef3b39b.json',
    scopes=scopes
)
gc = gspread.authorize(creds)

SHEET_ID = '1DTIRNy10Tx3GfhuFq46Eu2_4J74Z3ZiIh7ymZtetZUI'
sh = gc.open_by_key(SHEET_ID)

# --- Load Excel ---
wb = openpyxl.load_workbook('每日入院名單.xlsx')

def argb_to_rgb(argb_hex):
    """Convert openpyxl ARGB hex (e.g., 'FFBDD7EE') to Google Sheets RGB dict"""
    if not argb_hex or argb_hex == '00000000':
        return None
    s = str(argb_hex)
    if len(s) == 8:
        r, g, b = int(s[2:4], 16), int(s[4:6], 16), int(s[6:8], 16)
    elif len(s) == 6:
        r, g, b = int(s[0:2], 16), int(s[2:4], 16), int(s[4:6], 16)
    else:
        return None
    return {"red": r/255, "green": g/255, "blue": b/255}

def get_cell_data(ws, max_row, max_col):
    """Read all cell values from Excel worksheet"""
    data = []
    for r in range(1, max_row + 1):
        row = []
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                row.append('')
            else:
                row.append(str(v))
        data.append(row)
    return data

def get_actual_dimensions(ws):
    """Get actual data dimensions (not max_row which can be inflated)"""
    max_r, max_c = 0, 0
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=r, column=c).value is not None:
                max_r = max(max_r, r)
                max_c = max(max_c, c)
    return max(max_r, 1), max(max_c, 1)

def build_format_requests(ws, max_row, max_col, sheet_id):
    """Build batch format requests for Google Sheets API"""
    requests = []

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell_format = {}
            fields = []

            # Background color
            if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                rgb = argb_to_rgb(cell.fill.start_color.rgb)
                if rgb:
                    cell_format["backgroundColor"] = rgb
                    fields.append("userEnteredFormat.backgroundColor")

            # Font
            text_format = {}
            if cell.font:
                if cell.font.bold:
                    text_format["bold"] = True
                    fields.append("userEnteredFormat.textFormat.bold")
                if cell.font.size:
                    text_format["fontSize"] = cell.font.size
                    fields.append("userEnteredFormat.textFormat.fontSize")
                if cell.font.color and cell.font.color.rgb:
                    rgb = argb_to_rgb(cell.font.color.rgb)
                    if rgb:
                        text_format["foregroundColor"] = rgb
                        fields.append("userEnteredFormat.textFormat.foregroundColor")

            if text_format:
                cell_format["textFormat"] = text_format

            # Alignment
            if cell.alignment:
                ha = cell.alignment.horizontal
                va = cell.alignment.vertical
                wrap = cell.alignment.wrap_text
                if ha:
                    ha_map = {'left': 'LEFT', 'center': 'CENTER', 'right': 'RIGHT'}
                    if ha in ha_map:
                        cell_format["horizontalAlignment"] = ha_map[ha]
                        fields.append("userEnteredFormat.horizontalAlignment")
                if va:
                    va_map = {'top': 'TOP', 'center': 'MIDDLE', 'bottom': 'BOTTOM'}
                    if va in va_map:
                        cell_format["verticalAlignment"] = va_map[va]
                        fields.append("userEnteredFormat.verticalAlignment")
                if wrap:
                    cell_format["wrapStrategy"] = "WRAP"
                    fields.append("userEnteredFormat.wrapStrategy")

            if fields and cell_format:
                requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": sheet_id,
                            "startRowIndex": r - 1,
                            "endRowIndex": r,
                            "startColumnIndex": c - 1,
                            "endColumnIndex": c
                        },
                        "cell": {"userEnteredFormat": cell_format},
                        "fields": ",".join(fields)
                    }
                })

    return requests

def build_merge_requests(ws, sheet_id):
    """Build merge cell requests"""
    requests = []
    for mg in ws.merged_cells.ranges:
        requests.append({
            "mergeCells": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": mg.min_row - 1,
                    "endRowIndex": mg.max_row,
                    "startColumnIndex": mg.min_col - 1,
                    "endColumnIndex": mg.max_col
                },
                "mergeType": "MERGE_ALL"
            }
        })
    return requests

def build_column_width_requests(ws, max_col, sheet_id):
    """Build column width requests"""
    requests = []
    for c in range(1, max_col + 1):
        col_letter = openpyxl.utils.get_column_letter(c)
        dim = ws.column_dimensions.get(col_letter)
        if dim and dim.width:
            # Excel width to pixels (approximate: 1 Excel unit ≈ 7.5 pixels)
            pixel_width = int(dim.width * 7.5)
            requests.append({
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": sheet_id,
                        "dimension": "COLUMNS",
                        "startIndex": c - 1,
                        "endIndex": c
                    },
                    "properties": {"pixelSize": max(pixel_width, 30)},
                    "fields": "pixelSize"
                }
            })
    return requests

def build_note_requests(ws, max_row, max_col, sheet_id):
    """Build cell note (comment) requests"""
    requests = []
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if cell.comment and cell.comment.text:
                requests.append({
                    "updateCells": {
                        "rows": [{"values": [{"note": cell.comment.text[:5000]}]}],
                        "range": {
                            "sheetId": sheet_id,
                            "startRowIndex": r - 1,
                            "endRowIndex": r,
                            "startColumnIndex": c - 1,
                            "endColumnIndex": c
                        },
                        "fields": "note"
                    }
                })
    return requests

# --- Main migration ---
sheet_names = wb.sheetnames
print(f"開始遷移 {len(sheet_names)} 個工作表...")

# Rename default sheet to first Excel sheet name
existing_ws = sh.worksheets()
first_ws = existing_ws[0]

for idx, name in enumerate(sheet_names):
    ws_excel = wb[name]
    max_r, max_c = get_actual_dimensions(ws_excel)
    print(f"\n--- {name} ({max_r}x{max_c}) ---")

    # Create or reuse worksheet
    if idx == 0:
        gws = first_ws
        gws.update_title(name)
        gws.resize(rows=max(max_r, 1), cols=max(max_c, 1))
    else:
        gws = sh.add_worksheet(title=name, rows=max(max_r, 1), cols=max(max_c, 1))

    time.sleep(1)  # Rate limit

    # Write data
    data = get_cell_data(ws_excel, max_r, max_c)
    if data and data[0]:
        # Use batch update for data
        range_str = f"A1:{openpyxl.utils.get_column_letter(max_c)}{max_r}"
        gws.update(range_str, data, value_input_option='RAW')
        print(f"  資料寫入完成")

    time.sleep(1)

    # Build all format requests
    gs_id = gws.id
    all_requests = []

    # Column widths
    all_requests.extend(build_column_width_requests(ws_excel, max_c, gs_id))

    # Merged cells
    all_requests.extend(build_merge_requests(ws_excel, gs_id))

    # Cell formatting (batch in chunks to avoid API limits)
    fmt_requests = build_format_requests(ws_excel, max_r, max_c, gs_id)
    all_requests.extend(fmt_requests)

    # Cell notes/comments
    note_requests = build_note_requests(ws_excel, max_r, max_c, gs_id)
    all_requests.extend(note_requests)

    # Execute in batches of 500
    if all_requests:
        batch_size = 500
        for i in range(0, len(all_requests), batch_size):
            batch = all_requests[i:i+batch_size]
            sh.batch_update({"requests": batch})
            print(f"  格式批次 {i//batch_size + 1}/{(len(all_requests)-1)//batch_size + 1} 完成 ({len(batch)} requests)")
            time.sleep(2)

    print(f"  ✅ {name} 完成")

print(f"\n🎉 全部遷移完成！共 {len(sheet_names)} 個工作表")
print(f"Google Sheet: https://docs.google.com/spreadsheets/d/{SHEET_ID}")
